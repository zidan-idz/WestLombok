import openpyxl
from django.core.management.base import BaseCommand
from apps.core.models import Destination, District, Category
from django.utils import timezone
import os

class Command(BaseCommand):
    help = 'Exports Destination, District, and Category data to an Excel file.'

    def handle(self, *args, **kwargs):
        self.stdout.write(self.style.SUCCESS('Starting database export...'))

        # Create a workbook
        wb = openpyxl.Workbook()
        
        # --- Sheet 1: Destinations ---
        ws_dest = wb.active
        ws_dest.title = "Destinations"
        
        # Headers
        headers_dest = ['ID', 'Name', 'Slug', 'District', 'Category', 'Description', 'Views', 'Created At']
        ws_dest.append(headers_dest)
        
        # Data
        destinations = Destination.objects.all().select_related('district', 'category')
        for dest in destinations:
            ws_dest.append([
                dest.id,
                dest.name,
                dest.slug,
                dest.district.name if dest.district else "No District",
                dest.category.name if dest.category else "No Category",
                dest.description[:100] + "..." if dest.description else "", # Truncate for readability
                dest.view_count,
                dest.created_at.strftime('%Y-%m-%d %H:%M:%S') if dest.created_at else ""
            ])
        
        self.stdout.write(f"Exported {destinations.count()} destinations.")

        # --- Sheet 2: Districts ---
        ws_dist = wb.create_sheet(title="Districts")
        headers_dist = ['ID', 'Name', 'Slug', 'Description']
        ws_dist.append(headers_dist)
        
        districts = District.objects.all()
        for district in districts:
            ws_dist.append([
                district.id,
                district.name,
                district.slug,
                district.description[:100] + "..." if district.description else ""
            ])
            
        self.stdout.write(f"Exported {districts.count()} districts.")

        # --- Sheet 3: Categories ---
        ws_cat = wb.create_sheet(title="Categories")
        headers_cat = ['ID', 'Name', 'Slug', 'Icon', 'Description']
        ws_cat.append(headers_cat)
        
        categories = Category.objects.all()
        for cat in categories:
            ws_cat.append([
                cat.id,
                cat.name,
                cat.slug,
                cat.icon,
                cat.description[:100] + "..." if cat.description else ""
            ])
            
        self.stdout.write(f"Exported {categories.count()} categories.")

        # --- Save File ---
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f'West_Lombok_DB_Export_{timestamp}.xlsx'
        file_path = os.path.join(os.getcwd(), filename)
        
        wb.save(file_path)
        
        self.stdout.write(self.style.SUCCESS(f'Successfully exported database to: {file_path}'))
